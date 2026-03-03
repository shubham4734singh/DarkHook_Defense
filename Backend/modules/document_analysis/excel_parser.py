# ================================================================
# excel_parser.py — DarkHOOK_ Defence
# Version  : 2.0 — Enterprise Grade
# Purpose  : Excel file phishing detection using
#            16 industry-standard techniques
#
# Technique 1  -> File Type Validation
# Technique 2  -> Metadata Analysis
# Technique 3  -> Macro Detection
# Technique 4  -> Auto-Execution Detection
# Technique 5  -> VBA Behavior Analysis
# Technique 6  -> Macro Obfuscation Detection
# Technique 7  -> Embedded Object Analysis
# Technique 8  -> External Resource Analysis
# Technique 9  -> Content and Keyword Analysis
# Technique 10 -> URL Analysis
# Technique 11 -> Attack Chain Inference
# Technique 12 -> Heuristic Risk Scoring
# Technique 13 -> XLM Macro Detection
# Technique 14 -> Hidden Sheet Detection
# Technique 15 -> Advanced Formula Injection
# Technique 16 -> Power Query and Data Connection
#
# Libraries: openpyxl, oletools, zipfile,
#            hashlib, re, math
# ================================================================


# ----------------------------------------------------------------
# IMPORTS
# ----------------------------------------------------------------

import re
import os
import math
import hashlib
import zipfile
from urllib.parse import urlparse
from collections import Counter

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from oletools.olevba import VBA_Parser
    OLETOOLS_AVAILABLE = True
except ImportError:
    OLETOOLS_AVAILABLE = False


# ================================================================
# CONFIGURATION
# ================================================================

WEIGHTS = {
    # File structure
    "file_type_mismatch"         : 40,
    "corrupted_structure"        : 30,
    "double_extension"           : 35,
    "malformed_zip"              : 25,
    "xlsm_file"                  : 20,
    "xlsb_file"                  : 25,
    # Metadata
    "suspicious_metadata"        : 15,
    "wiped_metadata"             : 20,
    "external_template"          : 35,
    # Macro findings
    "malicious_macro"            : 40,
    "autoopen_macro"             : 35,
    "hidden_macro_stream"        : 35,
    # VBA behavior
    "suspicious_vba_api"         : 30,
    "powershell_in_vba"          : 40,
    "network_call_in_vba"        : 35,
    "file_system_access"         : 25,
    "registry_access"            : 30,
    "process_creation"           : 35,
    # Obfuscation
    "encoded_macro_payload"      : 35,
    "high_entropy_string"        : 25,
    "string_obfuscation"         : 25,
    # Embedded objects
    "embedded_ole_object"        : 30,
    "embedded_executable"        : 45,
    "embedded_script"            : 40,
    "double_extension_payload"   : 40,
    # External resources
    "dde_attack"                 : 45,
    "external_image_tracker"     : 20,
    "suspicious_relationship"    : 25,
    "hidden_relationship"        : 30,
    # Content
    "phishing_keyword"           : 10,
    "urgent_tone_detected"       : 15,
    "financial_terms_detected"   : 15,
    "credential_harvesting"      : 20,
    "enable_macro_lure"          : 35,
    "repeated_cta"               : 15,
    # URL findings
    "suspicious_url"             : 15,
    "ip_based_url"               : 30,
    "shortened_url"              : 20,
    "suspicious_tld"             : 20,
    "at_symbol_trick"            : 25,
    "mismatched_anchor"          : 25,
    # Attack chain
    "dropper_pattern"            : 40,
    "download_execute_pattern"   : 40,
    "multistage_indicator"       : 35,
    # Reputation
    "known_macro_signature"      : 45,
    # NEW — XLM macros
    "xlm_macro_detected"         : 40,
    "xlm_exec_command"           : 45,
    "xlm_run_command"            : 40,
    "xlm_call_command"           : 40,
    # NEW — Hidden sheets
    "hidden_sheet"               : 25,
    "very_hidden_sheet"          : 40,
    # NEW — Formula injection
    "formula_hyperlink_injection": 35,
    "webservice_formula"         : 45,
    "formula_obfuscation"        : 30,
    "char_concat_formula"        : 25,
    # NEW — Power Query
    "power_query_connection"     : 35,
    "suspicious_connection"      : 40,
    "ole_db_connection"          : 30,
    "external_data_connection"   : 30,
}


# ----------------------------------------------------------------
# PHISHING KEYWORDS — Technique 9
# ----------------------------------------------------------------

PHISHING_KEYWORDS = {

    "account_threats": [
        "verify your account",
        "confirm your account",
        "account suspended",
        "account will be closed",
        "account has been compromised",
        "account limited",
        "reactivate your account",
        "unlock your account",
        "account will be terminated",
        "unusual activity detected",
    ],

    "urgency_phrases": [
        "urgent action required",
        "immediate action required",
        "act now",
        "respond immediately",
        "expires today",
        "final warning",
        "last chance",
        "do not ignore",
        "failure to respond",
        "within 24 hours",
        "time sensitive",
    ],

    "credential_harvesting": [
        "enter your password",
        "confirm your password",
        "reset your password",
        "password expired",
        "enter your credentials",
        "verify your identity",
        "sign in to continue",
        "login credentials required",
        "enter your username",
        "enter your email",
    ],

    "financial_terms": [
        "bank account details",
        "credit card details",
        "debit card number",
        "billing information required",
        "payment details required",
        "update payment method",
        "transaction failed",
        "wire transfer",
        "gift card",
        "bitcoin payment",
        "refund pending",
    ],

    "reward_tricks": [
        "you have won",
        "prize money",
        "claim your reward",
        "lottery winner",
        "congratulations you won",
        "cash prize",
        "free gift",
    ],

    "legal_threats": [
        "legal action will be taken",
        "police complaint filed",
        "court notice",
        "government notice",
        "income tax department",
        "irs notice",
        "tax refund",
        "warrant issued",
        "arrest warrant",
    ],

    "india_specific": [
        "aadhar number",
        "pan card details",
        "kyc verification",
        "kyc update required",
        "upi details",
        "enter otp",
        "otp verification",
    ],

    "fake_security_alerts": [
        "your computer is infected",
        "virus detected",
        "security breach",
        "unauthorized access",
        "install this update",
        "download the security patch",
        "suspicious login attempt",
    ],

    "enable_macro_lures": [
        "enable macros to view",
        "enable editing to continue",
        "enable content to view",
        "click enable to continue",
        "must enable macros",
        "enable to view document",
        "protected document",
        "click enable content",
        "enable macros to see salary",
        "enable macros to view invoice",
    ],

    "download_tricks": [
        "click the link below",
        "click here to verify",
        "download the attachment",
        "view your invoice",
        "access your account here",
        "download now",
        "click here immediately",
    ],
}


# ----------------------------------------------------------------
# DANGEROUS VBA APIs — Technique 5
# ----------------------------------------------------------------

DANGEROUS_VBA_APIS = {

    "shell_execution": [
        "shell(",
        "shell ",
        "wscript.shell",
        "shell.application",
        "shellexecute",
        "winexec(",
        'createobject("wscript',
    ],

    "powershell": [
        "powershell",
        "powershell.exe",
        "-encodedcommand",
        "-enc ",
        "invoke-expression",
        "iex(",
        "invoke-webrequest",
    ],

    "network_calls": [
        "xmlhttp",
        "xmlhttprequest",
        "urldownloadtofile",
        "winhttprequest",
        "msxml2.xmlhttp",
        "http.open",
    ],

    "file_system": [
        "filesystemobject",
        "createtextfile",
        "opentextfile",
        "copyfile",
        "deletefile",
        "movefile",
        "kill(",
        "filecopy",
    ],

    "registry": [
        "regwrite",
        "regread",
        "regdelete",
        "hkey_local_machine",
        "hkey_current_user",
    ],

    "process_creation": [
        "createprocess",
        "getobject(",
        "wmi",
        "win32_process",
    ],
}


# AutoOpen names — Excel specific
AUTOOPEN_NAMES = [
    "auto_open",
    "workbook_open",
    "auto_close",
    "workbook_close",
    "workbook_activate",
    "worksheet_activate",
    "workbook_beforeclose",
    "auto_exec",
    "autoopen",
]


# DDE attack patterns
DDE_PATTERNS = [
    "=cmd|",
    "=powershell|",
    "=msexcel|",
    "=dde(",
    "=ddeauto(",
    "|'/c",
    "cmd.exe",
    "=rundll32",
    "=mshta",
]


# XLM dangerous commands — Technique 13
XLM_DANGEROUS_COMMANDS = [
    "exec(",
    "run(",
    "call(",
    "=exec",
    "=run(",
    "=call(",
    "halt()",
    "return(",
    "formula(",
    "set.value(",
    "files(",
    "fopen(",
    "fwrite(",
    "register(",
]


# Suspicious formula functions — Technique 15
SUSPICIOUS_FORMULAS = [
    "=hyperlink(",
    "=webservice(",
    "=indirect(",
    "=offset(",
    "=filterxml(",
    "=encodeurl(",
]


# Known malicious macro signatures
KNOWN_MACRO_SIGNATURES = [
    "powershell -nop -w hidden",
    "iex(new-object net.webclient",
    "urldownloadtofile",
    'shell("cmd /c',
    'wscript.shell").run',
    'createobject("adodb.stream")',
]


# URL shorteners
URL_SHORTENERS = [
    "bit.ly", "tinyurl.com", "t.co",
    "goo.gl", "ow.ly", "is.gd",
    "buff.ly", "rebrand.ly", "cutt.ly",
    "shorturl.at", "tiny.cc", "rb.gy",
]


# Suspicious TLDs
SUSPICIOUS_TLDS = [
    ".xyz", ".top", ".ru", ".tk",
    ".ml", ".ga", ".cf", ".gq",
    ".pw", ".click", ".download",
    ".loan", ".work", ".party",
]


# Safe Microsoft domains
SAFE_DOMAINS = [
    "schemas.openxmlformats.org",
    "schemas.microsoft.com",
    "purl.org",
    "www.w3.org",
    "microsoft.com",
    "office.com",
]


# ================================================================
# HELPER FUNCTIONS
# ================================================================

def calculate_entropy(text):
    """
    Shannon entropy measures randomness.
    Normal text  -> 3.0 to 4.5
    Encoded data -> 6.0 to 8.0
    """
    if not text or len(text) < 20:
        return 0.0
    counter = Counter(text)
    length  = len(text)
    entropy = -sum(
        (c / length) * math.log2(c / length)
        for c in counter.values()
    )
    return round(entropy, 2)


def is_ip_url(url):
    """Returns True if URL uses IP instead of domain"""
    try:
        host = urlparse(url).netloc
        if ":" in host:
            host = host.split(":")[0]
        return bool(re.match(
            r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', host
        ))
    except Exception:
        return False


def analyze_url(url):
    """Full URL analysis — 9 checks"""
    url_findings = []
    url_details  = []

    try:
        parsed = urlparse(url)
        domain = parsed.netloc.lower()
        path   = parsed.path.lower()

        if is_ip_url(url):
            url_findings.append("ip_based_url")
            url_details.append("IP-based URL: " + url)

        for s in URL_SHORTENERS:
            if s in domain:
                url_findings.append("shortened_url")
                url_details.append(
                    "URL shortener (" + s + "): " + url
                )
                break

        for tld in SUSPICIOUS_TLDS:
            if domain.endswith(tld):
                url_findings.append("suspicious_tld")
                url_details.append(
                    "Suspicious TLD (" + tld + "): " + url
                )
                break

        if "@" in url:
            url_findings.append("at_symbol_trick")
            url_details.append("@ redirect trick: " + url)

        if url.startswith("http://"):
            url_findings.append("suspicious_url")
            url_details.append("Insecure HTTP: " + url)

        if len(url) > 200:
            url_findings.append("suspicious_url")
            url_details.append(
                "Suspicious long URL: " + url[:60] + "..."
            )

        fake_words = [
            "login", "signin", "verify", "secure",
            "account", "update", "confirm", "banking",
            "password", "credential",
        ]
        for word in fake_words:
            if word in path:
                url_findings.append("suspicious_url")
                url_details.append(
                    "Login keyword in URL (" + word + "): " + url
                )
                break

        if re.search(
            r'[àáâãäåæçèéêëìíîïðñòóôõöøùúûü]', domain
        ):
            url_findings.append("suspicious_url")
            url_details.append("Homograph domain: " + domain)

        if url.count("http") > 1:
            url_findings.append("suspicious_url")
            url_details.append("Double HTTP obfuscation: " + url)

    except Exception as e:
        url_details.append("URL error: " + str(e))

    return url_findings, url_details


# ================================================================
# TECHNIQUE 1 — File Type and Structure Validation
# ================================================================

def technique1_file_validation(file_path):
    """
    Verifies file is actually a real Excel file.
    Checks extension, structure and dangerous formats.
    """
    findings = []
    details  = []

    details.append("--- TECHNIQUE 1: FILE VALIDATION ---")

    filename = os.path.basename(file_path)
    ext      = os.path.splitext(filename)[1].lower()

    valid_extensions = [
        ".xlsx", ".xlsm", ".xls",
        ".xlsb", ".xltx", ".xltm", ".csv",
    ]
    if ext not in valid_extensions:
        findings.append("file_type_mismatch")
        details.append("Invalid extension: " + ext)
    else:
        details.append("Extension valid: " + ext)

    if ext == ".xlsm":
        findings.append("xlsm_file")
        details.append(
            "WARNING: .xlsm = macro-enabled Excel!"
        )

    if ext == ".xlsb":
        findings.append("xlsb_file")
        details.append(
            "WARNING: .xlsb = binary format — harder to scan!"
        )

    if ext == ".xltm":
        findings.append("malicious_macro")
        details.append(
            "CRITICAL: .xltm = macro template — auto-runs on open!"
        )

    # Double extension check — ignores version numbers
    name_without_ext = os.path.splitext(filename)[0]
    if "." in name_without_ext:
        part_after_dot = name_without_ext.split(".")[-1].strip()
        dangerous_inner_exts = [
            "pdf", "exe", "dll", "bat", "cmd",
            "ps1", "vbs", "js", "hta", "scr",
            "zip", "rar", "7z", "iso",
        ]
        if part_after_dot.lower() in dangerous_inner_exts:
            findings.append("double_extension")
            details.append(
                "Double extension detected: " + filename
            )
        elif part_after_dot.isdigit():
            details.append(
                "Note: Dot in filename is version number — safe"
            )
        else:
            details.append(
                "Note: Dot in filename — verify manually"
            )

    # ZIP structure check
    if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        try:
            with zipfile.ZipFile(file_path, "r") as z:
                names = z.namelist()
                if "xl/workbook.xml" not in names:
                    findings.append("corrupted_structure")
                    details.append(
                        "Malformed Excel: xl/workbook.xml missing"
                    )
                else:
                    details.append("Excel ZIP structure valid")
        except zipfile.BadZipFile:
            findings.append("malformed_zip")
            details.append(
                "File is NOT valid ZIP — possible disguised malware!"
            )
        except Exception as e:
            details.append("ZIP check error: " + str(e))

    details.append("Technique 1 findings: " + str(len(findings)))
    return findings, details


# ================================================================
# TECHNIQUE 2 — Metadata Analysis
# ================================================================

def technique2_metadata(file_path):
    """
    Analyses hidden metadata inside Excel file.
    """
    findings = []
    details  = []

    details.append("")
    details.append("--- TECHNIQUE 2: METADATA ANALYSIS ---")

    if not OPENPYXL_AVAILABLE:
        details.append("openpyxl not available — skipping")
        return findings, details

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        details.append("Skipped for " + ext + " format")
        return findings, details

    try:
        wb    = openpyxl.load_workbook(file_path, read_only=True)
        props = wb.properties

        author   = props.creator  or ""
        created  = props.created
        modified = props.modified
        revision = props.revision or 0

        details.append("Author   : " + (author or "EMPTY"))
        details.append("Created  : " + str(created))
        details.append("Modified : " + str(modified))
        details.append("Revision : " + str(revision))

        wb.close()

        if not author or author.strip() == "":
            findings.append("wiped_metadata")
            details.append(
                "Author field empty — metadata may be wiped!"
            )

        suspicious_names = [
            "admin", "user", "test", "administrator",
            "temp", "unknown", "abc", "owner",
        ]
        if author.lower().strip() in suspicious_names:
            findings.append("suspicious_metadata")
            details.append(
                "Suspicious author name: " + author
            )

        if created and modified:
            diff = abs((modified - created).total_seconds())
            if diff < 2:
                findings.append("suspicious_metadata")
                details.append(
                    "Created and modified at same time — "
                    "may be auto-generated"
                )

        try:
            if int(revision) == 1:
                findings.append("suspicious_metadata")
                details.append(
                    "Revision = 1 — may be auto-generated"
                )
        except Exception:
            pass

    except Exception as e:
        details.append("Metadata error: " + str(e))

    details.append("Technique 2 findings: " + str(len(findings)))
    return findings, details


# ================================================================
# TECHNIQUES 3, 4, 5, 6 — Macro Analysis
# ================================================================

def techniques3456_macro_analysis(file_path):
    """
    Technique 3 -> Macro presence
    Technique 4 -> Auto-execution
    Technique 5 -> VBA behavior
    Technique 6 -> Obfuscation
    """
    findings = []
    details  = []

    details.append("")
    details.append("--- TECHNIQUES 3-6: MACRO ANALYSIS ---")

    if not OLETOOLS_AVAILABLE:
        details.append("oletools not available — skipping")
        return findings, details

    try:
        vba = VBA_Parser(file_path)

        if not vba.detect_vba_macros():
            details.append("Technique 3: No VBA macros detected")
            vba.close()
            return findings, details

        findings.append("malicious_macro")
        details.append("Technique 3: VBA macros DETECTED!")

        try:
            results = vba.analyze_macros()
            for kw_type, kw_keyword, kw_description in results:
                if "suspicious" in kw_type.lower():
                    findings.append("hidden_macro_stream")
                    details.append(
                        "Hidden macro stream: " + str(kw_description)
                    )
        except Exception:
            pass

        all_vba_code = ""
        for (filename, stream_path,
             vba_filename, vba_code) in vba.extract_macros():
            all_vba_code += vba_code.lower() + "\n"

        # Technique 4 — Auto execution
        for auto_name in AUTOOPEN_NAMES:
            if auto_name in all_vba_code:
                findings.append("autoopen_macro")
                details.append(
                    "Technique 4: AutoOpen: '" +
                    auto_name + "' — RUNS ON FILE OPEN!"
                )

        # Technique 5 — VBA behavior
        for category, apis in DANGEROUS_VBA_APIS.items():
            for api in apis:
                if api in all_vba_code:
                    if category == "shell_execution":
                        findings.append("suspicious_vba_api")
                        details.append(
                            "Technique 5 [Shell]: " + api
                        )
                    elif category == "powershell":
                        findings.append("powershell_in_vba")
                        details.append(
                            "Technique 5 [PowerShell]: " + api
                        )
                    elif category == "network_calls":
                        findings.append("network_call_in_vba")
                        details.append(
                            "Technique 5 [Network]: " + api
                        )
                    elif category == "file_system":
                        findings.append("file_system_access")
                        details.append(
                            "Technique 5 [FileSystem]: " + api
                        )
                    elif category == "registry":
                        findings.append("registry_access")
                        details.append(
                            "Technique 5 [Registry]: " + api
                        )
                    elif category == "process_creation":
                        findings.append("process_creation")
                        details.append(
                            "Technique 5 [Process]: " + api
                        )

        # Technique 6 — Obfuscation
        b64_pattern = re.compile(r'[A-Za-z0-9+/]{50,}={0,2}')
        b64_matches = b64_pattern.findall(all_vba_code)
        if b64_matches:
            findings.append("encoded_macro_payload")
            details.append(
                "Technique 6 [Base64]: " +
                b64_matches[0][:50] + "..."
            )

        hex_pattern = re.compile(r'[0-9a-f]{40,}')
        hex_matches = hex_pattern.findall(all_vba_code)
        if hex_matches:
            findings.append("encoded_macro_payload")
            details.append(
                "Technique 6 [Hex]: " +
                hex_matches[0][:50] + "..."
            )

        concat_patterns = [
            r'"\s*&\s*"',
            r'chr\(\d+\)',
            r'chrw\(\d+\)',
        ]
        for pattern in concat_patterns:
            if re.search(pattern, all_vba_code):
                findings.append("string_obfuscation")
                details.append(
                    "Technique 6 [Obfuscation]: " + pattern
                )

        entropy = calculate_entropy(all_vba_code[:1000])
        if entropy > 5.5:
            findings.append("high_entropy_string")
            details.append(
                "Technique 6 [Entropy]: " +
                str(entropy) + " — obfuscated"
            )

        for sig in KNOWN_MACRO_SIGNATURES:
            if sig in all_vba_code:
                findings.append("known_macro_signature")
                details.append(
                    "Known malicious signature: " + sig[:50]
                )

        vba.close()

    except Exception as e:
        details.append("Macro analysis error: " + str(e))

    details.append(
        "Techniques 3-6 findings: " + str(len(findings))
    )
    return findings, details


# ================================================================
# TECHNIQUE 7 — Embedded Object Analysis
# ================================================================

def technique7_embedded_objects(file_path):
    """
    Checks for hidden files embedded inside Excel.
    """
    findings = []
    details  = []

    details.append("")
    details.append("--- TECHNIQUE 7: EMBEDDED OBJECT ANALYSIS ---")

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        details.append("Skipped for " + ext + " format")
        return findings, details

    try:
        with zipfile.ZipFile(file_path, "r") as z:
            all_files = z.namelist()

            ole_files = [
                f for f in all_files
                if "embeddings" in f.lower()
                or "oleobject" in f.lower()
                or "drawings" in f.lower()
            ]

            if ole_files:
                for ole_file in ole_files:
                    findings.append("embedded_ole_object")
                    details.append(
                        "Embedded OLE object: " + ole_file
                    )
                    dangerous_exts = [
                        ".exe", ".dll", ".bat", ".cmd",
                        ".ps1", ".vbs", ".js", ".hta",
                        ".scr", ".pif", ".com",
                    ]
                    for dext in dangerous_exts:
                        if ole_file.lower().endswith(dext):
                            findings.append("embedded_executable")
                            details.append(
                                "CRITICAL: Embedded executable: " +
                                ole_file
                            )
                    base = os.path.splitext(ole_file)[0]
                    if "." in os.path.basename(base):
                        findings.append("double_extension_payload")
                        details.append(
                            "Double extension in embedded: " + ole_file
                        )

            script_files = [
                f for f in all_files
                if any(
                    f.lower().endswith(e)
                    for e in [
                        ".js", ".vbs", ".ps1",
                        ".bat", ".cmd", ".hta",
                    ]
                )
            ]
            for sf in script_files:
                findings.append("embedded_script")
                details.append(
                    "CRITICAL: Embedded script: " + sf
                )

            for fname in all_files:
                try:
                    data = z.read(fname)
                    if data[:2] == b"MZ":
                        findings.append("embedded_executable")
                        details.append(
                            "CRITICAL: EXE signature (MZ) in: " +
                            fname
                        )
                except Exception:
                    pass

            if not ole_files and not script_files:
                details.append("No suspicious embedded objects")

    except Exception as e:
        details.append("Embedded object error: " + str(e))

    details.append("Technique 7 findings: " + str(len(findings)))
    return findings, details


# ================================================================
# TECHNIQUE 8 — External Resource Analysis
# ================================================================

def technique8_external_resources(file_path):
    """
    Checks for DDE attacks and external connections.
    """
    findings = []
    details  = []

    details.append("")
    details.append(
        "--- TECHNIQUE 8: EXTERNAL RESOURCE ANALYSIS ---"
    )

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        details.append("Skipped for " + ext + " format")
        return findings, details

    try:
        with zipfile.ZipFile(file_path, "r") as z:

            rel_files = [
                f for f in z.namelist()
                if f.endswith(".rels")
            ]

            for rel_file in rel_files:
                with z.open(rel_file) as f:
                    content = f.read().decode(
                        "utf-8", errors="ignore"
                    )

                    if "attachedTemplate" in content:
                        url_match = re.search(
                            r'Target="(https?://[^"]+)"', content
                        )
                        if url_match:
                            findings.append("external_template")
                            details.append(
                                "CRITICAL: External template: " +
                                url_match.group(1)
                            )

                    if "image" in rel_file.lower():
                        ext_urls = re.findall(
                            r'Target="(https?://[^"]+)"', content
                        )
                        for url in ext_urls:
                            findings.append("external_image_tracker")
                            details.append(
                                "External image tracker: " + url
                            )

                    all_ext = re.findall(
                        r'TargetMode="External"[^>]*'
                        r'Target="([^"]+)"',
                        content
                    )
                    for target in all_ext:
                        if target.startswith("http"):
                            findings.append(
                                "suspicious_relationship"
                            )
                            details.append(
                                "External relationship: " + target
                            )

            # Check XML files for DDE
            xml_files = [
                f for f in z.namelist()
                if f.endswith(".xml")
            ]
            for xml_file in xml_files:
                try:
                    with z.open(xml_file) as f:
                        content = f.read().decode(
                            "utf-8", errors="ignore"
                        ).lower()
                        for dde in DDE_PATTERNS:
                            if dde.lower() in content:
                                findings.append("dde_attack")
                                details.append(
                                    "CRITICAL: DDE pattern in " +
                                    xml_file + ": " + dde
                                )
                                break
                except Exception:
                    pass

            # Check settings — ignore safe Microsoft URLs
            settings_files = [
                f for f in z.namelist()
                if "settings" in f.lower()
                and f.endswith(".xml")
            ]
            for settings_file in settings_files:
                try:
                    with z.open(settings_file) as f:
                        settings = f.read().decode(
                            "utf-8", errors="ignore"
                        )
                        if ("http://" in settings or
                                "https://" in settings):
                            urls_found = re.findall(
                                r'https?://[^\s"<>]+', settings
                            )
                            for url in urls_found:
                                is_safe = any(
                                    s in url.lower()
                                    for s in SAFE_DOMAINS
                                )
                                if not is_safe:
                                    findings.append(
                                        "hidden_relationship"
                                    )
                                    details.append(
                                        "Suspicious URL in "
                                        "settings: " + url
                                    )
                except Exception:
                    pass

    except Exception as e:
        details.append("External resource error: " + str(e))

    if not findings:
        details.append("No external resources detected")

    details.append("Technique 8 findings: " + str(len(findings)))
    return findings, details


# ================================================================
# TECHNIQUES 9 and 10 — Content and URL Analysis
# ================================================================

def techniques910_content_url(file_path):
    """
    Technique 9  -> Phishing keyword analysis
    Technique 10 -> URL analysis
    """
    findings = []
    details  = []

    details.append("")
    details.append(
        "--- TECHNIQUES 9-10: CONTENT AND URL ANALYSIS ---"
    )

    if not OPENPYXL_AVAILABLE:
        details.append("openpyxl not available — skipping")
        return findings, details

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        details.append("Skipped for " + ext + " format")
        return findings, details

    try:
        wb        = openpyxl.load_workbook(
            file_path, read_only=True
        )
        full_text = ""
        all_urls  = []

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_str = str(cell.value).lower()
                        full_text += cell_str + " "

                        # Check DDE in formulas
                        if cell_str.startswith("="):
                            for dde in DDE_PATTERNS:
                                if dde.lower() in cell_str:
                                    findings.append("dde_attack")
                                    details.append(
                                        "CRITICAL: DDE formula: " +
                                        cell_str[:80]
                                    )
                                    break

                        # Extract URLs
                        url_pattern = re.compile(
                            r'https?://[^\s<>"{}|\\^`\[\]]+'
                        )
                        for url in url_pattern.findall(cell_str):
                            all_urls.append(("cell", url, ""))

                    # Extract hyperlinks
                    if cell.hyperlink:
                        try:
                            url = str(
                                cell.hyperlink.target or ""
                            )
                            if url.startswith("http"):
                                visible = str(cell.value or "")
                                all_urls.append(
                                    ("hyperlink", url, visible)
                                )
                        except Exception:
                            pass

        wb.close()

        # Technique 10 — URL Analysis
        details.append(
            "Total URLs found: " + str(len(all_urls))
        )

        if len(all_urls) > 10:
            findings.append("suspicious_url")
            details.append(
                "High URL count: " +
                str(len(all_urls)) + " — suspicious"
            )

        for url_type, url, visible_text in all_urls:
            url_findings, url_details = analyze_url(url)
            if url_findings:
                findings.extend(url_findings)
                for d in url_details:
                    details.append(
                        "  [" + url_type + "] " + d
                    )
            else:
                details.append(
                    "  [" + url_type + "] Link: " + url
                )

            if visible_text and url:
                vis_lower = visible_text.lower()
                url_lower = url.lower()
                if ("http" in vis_lower and
                        vis_lower.strip() != url_lower.strip()):
                    findings.append("mismatched_anchor")
                    details.append(
                        "Mismatched anchor: '" +
                        visible_text[:40] + "' -> '" +
                        url[:40] + "'"
                    )

        # Technique 9 — Keyword Analysis
        urgency_count    = 0
        financial_count  = 0
        credential_count = 0
        enable_count     = 0
        cta_count        = 0

        for category, keywords in PHISHING_KEYWORDS.items():
            for keyword in keywords:
                count = full_text.count(keyword)
                if count > 0:
                    findings.append("phishing_keyword")
                    details.append(
                        "[" + category + "] '" +
                        keyword + "' x" + str(count)
                    )
                    if category == "urgency_phrases":
                        urgency_count += count
                    elif category == "financial_terms":
                        financial_count += count
                    elif category == "credential_harvesting":
                        credential_count += count
                    elif category == "enable_macro_lures":
                        enable_count += count
                    elif category == "download_tricks":
                        cta_count += count

        if urgency_count >= 2:
            findings.append("urgent_tone_detected")
            details.append(
                "Urgency tone: " + str(urgency_count) + " phrases"
            )
        if financial_count >= 2:
            findings.append("financial_terms_detected")
            details.append(
                "Financial targeting: " +
                str(financial_count) + " terms"
            )
        if credential_count >= 1:
            findings.append("credential_harvesting")
            details.append(
                "Credential harvesting: " +
                str(credential_count) + " phrases"
            )
        if enable_count >= 1:
            findings.append("enable_macro_lure")
            details.append(
                "CRITICAL: Enable macro lure: " +
                str(enable_count) + "x detected!"
            )
        if cta_count >= 3:
            findings.append("repeated_cta")
            details.append(
                "Repeated CTA: " + str(cta_count) + "x"
            )

    except Exception as e:
        details.append("Content/URL error: " + str(e))

    details.append(
        "Techniques 9-10 findings: " + str(len(findings))
    )
    return findings, details


# ================================================================
# TECHNIQUE 11 — Attack Chain Inference
# ================================================================

def technique11_attack_chain(all_findings):
    """
    Infers complete attack sequence from all findings.
    """
    findings = []
    details  = []

    details.append("")
    details.append(
        "--- TECHNIQUE 11: ATTACK CHAIN INFERENCE ---"
    )

    fs = set(all_findings)

    if "malicious_macro" in fs and "network_call_in_vba" in fs:
        findings.append("download_execute_pattern")
        details.append(
            "ATTACK CHAIN: Macro -> Network -> Download -> Execute"
        )

    if "dde_attack" in fs:
        findings.append("multistage_indicator")
        details.append(
            "ATTACK CHAIN: DDE formula -> User clicks update "
            "-> Command executes"
        )

    if "enable_macro_lure" in fs and "malicious_macro" in fs:
        findings.append("dropper_pattern")
        details.append(
            "ATTACK CHAIN: Blurred content -> "
            "User enables macros -> Payload executes"
        )

    if "credential_harvesting" in fs and (
        "suspicious_url" in fs or "ip_based_url" in fs
    ):
        findings.append("dropper_pattern")
        details.append(
            "ATTACK CHAIN: Credential lure -> "
            "Fake login -> Steal credentials"
        )

    if ("embedded_executable" in fs or
            "embedded_ole_object" in fs):
        findings.append("dropper_pattern")
        details.append(
            "ATTACK CHAIN: Embedded payload -> "
            "File opened -> EXE extracted and run"
        )

    if "external_template" in fs:
        findings.append("multistage_indicator")
        details.append(
            "ATTACK CHAIN: External template loads -> "
            "Remote macro executes"
        )

    if "xlm_exec_command" in fs or "xlm_run_command" in fs:
        findings.append("dropper_pattern")
        details.append(
            "ATTACK CHAIN: XLM macro -> EXEC/RUN command -> "
            "Silent execution on open"
        )

    if "power_query_connection" in fs:
        findings.append("multistage_indicator")
        details.append(
            "ATTACK CHAIN: Power Query -> "
            "Fetch remote payload -> Execute"
        )

    if "webservice_formula" in fs:
        findings.append("multistage_indicator")
        details.append(
            "ATTACK CHAIN: WEBSERVICE formula -> "
            "Send data to attacker server"
        )

    if not findings:
        details.append("No clear attack chain identified")

    details.append(
        "Technique 11 findings: " + str(len(findings))
    )
    return findings, details


# ================================================================
# TECHNIQUE 12 — Heuristic Risk Scoring
# ================================================================

def technique12_scoring(all_findings):
    """
    Converts all findings into final weighted score.
    """
    total_score = 0
    breakdown   = {}

    for finding in all_findings:
        weight = WEIGHTS.get(finding, 5)
        total_score += weight

        if finding in breakdown:
            breakdown[finding]["count"] += 1
            breakdown[finding]["score"] += weight
        else:
            breakdown[finding] = {
                "count" : 1,
                "score" : weight,
            }

    total_score = min(total_score, 100)

    if total_score < 30:
        verdict  = "Low Risk"
        severity = "LOW"
    elif total_score < 60:
        verdict  = "Medium Risk"
        severity = "MEDIUM"
    elif total_score < 80:
        verdict  = "High Risk"
        severity = "HIGH"
    else:
        verdict  = "Critical — Likely Phishing"
        severity = "CRITICAL"

    return total_score, verdict, severity, breakdown


# ================================================================
# TECHNIQUE 13 — XLM Macro Detection
# ================================================================

def technique13_xlm_macros(file_path):
    """
    Detects Excel 4.0 (XLM) macros.
    XLM macros bypass many VBA-only scanners.
    Very old but still actively abused by attackers.
    """
    findings = []
    details  = []

    details.append("")
    details.append("--- TECHNIQUE 13: XLM MACRO DETECTION ---")

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in [".xlsx", ".xlsm", ".xls", ".xlsb", ".xltm"]:
        details.append("Skipped for " + ext + " format")
        return findings, details

    try:
        with zipfile.ZipFile(file_path, "r") as z:
            xml_files = [
                f for f in z.namelist()
                if f.endswith(".xml")
            ]

            for xml_file in xml_files:
                try:
                    with z.open(xml_file) as f:
                        content = f.read().decode(
                            "utf-8", errors="ignore"
                        ).lower()

                        # XLM macro sheet indicator
                        if "macrosheets" in content:
                            findings.append("xlm_macro_detected")
                            details.append(
                                "CRITICAL: XLM macro sheet "
                                "detected in: " + xml_file
                            )

                        # Check for dangerous XLM commands
                        for cmd in XLM_DANGEROUS_COMMANDS:
                            if cmd.lower() in content:
                                if "exec" in cmd:
                                    findings.append(
                                        "xlm_exec_command"
                                    )
                                    details.append(
                                        "CRITICAL: XLM EXEC "
                                        "command found!"
                                    )
                                elif "run" in cmd:
                                    findings.append(
                                        "xlm_run_command"
                                    )
                                    details.append(
                                        "CRITICAL: XLM RUN "
                                        "command found!"
                                    )
                                elif "call" in cmd:
                                    findings.append(
                                        "xlm_call_command"
                                    )
                                    details.append(
                                        "CRITICAL: XLM CALL "
                                        "command found!"
                                    )
                                else:
                                    findings.append(
                                        "xlm_macro_detected"
                                    )
                                    details.append(
                                        "XLM command found: " + cmd
                                    )

                except Exception:
                    pass

        # Also check using oletools if available
        if OLETOOLS_AVAILABLE:
            try:
                vba = VBA_Parser(file_path)
                # oletools can detect XLM sheets
                if hasattr(vba, 'xlm_macros'):
                    xlm = vba.xlm_macros
                    if xlm:
                        findings.append("xlm_macro_detected")
                        details.append(
                            "oletools: XLM macros confirmed!"
                        )
                vba.close()
            except Exception:
                pass

        if not findings:
            details.append("No XLM macros detected")

    except Exception as e:
        details.append("XLM check error: " + str(e))

    details.append("Technique 13 findings: " + str(len(findings)))
    return findings, details


# ================================================================
# TECHNIQUE 14 — Hidden Sheet Detection
# ================================================================

def technique14_hidden_sheets(file_path):
    """
    Detects hidden and very hidden sheets.
    Attackers hide malicious content in invisible sheets.

    Sheet visibility types:
    visible    = normal sheet everyone can see
    hidden     = hidden but can be unhidden from menu
    veryHidden = CANNOT be unhidden normally — very suspicious!
    """
    findings = []
    details  = []

    details.append("")
    details.append("--- TECHNIQUE 14: HIDDEN SHEET DETECTION ---")

    if not OPENPYXL_AVAILABLE:
        details.append("openpyxl not available — skipping")
        return findings, details

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        details.append("Skipped for " + ext + " format")
        return findings, details

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            state = sheet.sheet_state

            if state == "hidden":
                findings.append("hidden_sheet")
                details.append(
                    "Hidden sheet found: '" + sheet_name +
                    "' — check for hidden content"
                )

            elif state == "veryHidden":
                findings.append("very_hidden_sheet")
                details.append(
                    "CRITICAL: Very hidden sheet: '" +
                    sheet_name +
                    "' — cannot be unhidden normally!"
                )

            else:
                details.append(
                    "Sheet visible: '" + sheet_name + "'"
                )

        wb.close()

        # Also check workbook XML directly for hidden sheets
        with zipfile.ZipFile(file_path, "r") as z:
            if "xl/workbook.xml" in z.namelist():
                with z.open("xl/workbook.xml") as f:
                    content = f.read().decode(
                        "utf-8", errors="ignore"
                    )

                    # Count veryHidden occurrences
                    very_hidden_count = content.count(
                        'state="veryHidden"'
                    )
                    if very_hidden_count > 0:
                        findings.append("very_hidden_sheet")
                        details.append(
                            "CRITICAL: " +
                            str(very_hidden_count) +
                            " veryHidden sheet(s) in workbook.xml!"
                        )

        if not findings:
            details.append("No hidden sheets detected")

    except Exception as e:
        details.append("Hidden sheet error: " + str(e))

    details.append("Technique 14 findings: " + str(len(findings)))
    return findings, details


# ================================================================
# TECHNIQUE 15 — Advanced Formula Injection
# ================================================================

def technique15_formula_injection(file_path):
    """
    Detects weaponized Excel formulas beyond DDE.

    Dangerous formulas:
    =HYPERLINK() -> opens malicious URL on click
    =WEBSERVICE() -> sends data to attacker server
    =CHAR()+concat -> builds hidden commands character by character
    =INDIRECT()   -> dynamic cell reference obfuscation
    """
    findings = []
    details  = []

    details.append("")
    details.append(
        "--- TECHNIQUE 15: ADVANCED FORMULA INJECTION ---"
    )

    if not OPENPYXL_AVAILABLE:
        details.append("openpyxl not available — skipping")
        return findings, details

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        details.append("Skipped for " + ext + " format")
        return findings, details

    try:
        wb = openpyxl.load_workbook(
            file_path, read_only=True
        )

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_str = str(cell.value).lower()

                        # Only check formula cells
                        if not cell_str.startswith("="):
                            continue

                        # Check suspicious formulas
                        for formula in SUSPICIOUS_FORMULAS:
                            if formula in cell_str:

                                if "hyperlink" in formula:
                                    # Check if URL inside is suspicious
                                    url_match = re.search(
                                        r'https?://[^\s")\]]+',
                                        cell_str
                                    )
                                    if url_match:
                                        url = url_match.group(0)
                                        url_f, _ = analyze_url(url)
                                        if url_f:
                                            findings.append(
                                                "formula_hyperlink_injection"
                                            )
                                            details.append(
                                                "CRITICAL: Malicious "
                                                "HYPERLINK formula: " +
                                                cell_str[:80]
                                            )

                                elif "webservice" in formula:
                                    findings.append(
                                        "webservice_formula"
                                    )
                                    details.append(
                                        "CRITICAL: WEBSERVICE formula "
                                        "— sends data to external "
                                        "server: " + cell_str[:80]
                                    )

                                elif "indirect" in formula:
                                    findings.append(
                                        "formula_obfuscation"
                                    )
                                    details.append(
                                        "INDIRECT formula detected "
                                        "— possible obfuscation: " +
                                        cell_str[:80]
                                    )

                        # Check CHAR() concatenation obfuscation
                        # Example: =CHAR(112)&CHAR(111)&CHAR(119)
                        # Builds "pow" → part of "powershell"
                        char_count = cell_str.count("char(")
                        if char_count >= 4:
                            findings.append("char_concat_formula")
                            details.append(
                                "CRITICAL: CHAR() concatenation "
                                "obfuscation (" +
                                str(char_count) +
                                " CHAR calls) — hiding command!"
                            )

        wb.close()

        if not findings:
            details.append("No formula injection detected")

    except Exception as e:
        details.append("Formula injection error: " + str(e))

    details.append("Technique 15 findings: " + str(len(findings)))
    return findings, details


# ================================================================
# TECHNIQUE 16 — Power Query and Data Connection Analysis
# ================================================================

def technique16_power_query(file_path):
    """
    Detects suspicious external data connections.

    Power Query / OLE DB connections can:
    - Pull remote payloads silently
    - Exfiltrate data to attacker server
    - Execute commands via connection strings
    """
    findings = []
    details  = []

    details.append("")
    details.append(
        "--- TECHNIQUE 16: POWER QUERY AND DATA CONNECTIONS ---"
    )

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        details.append("Skipped for " + ext + " format")
        return findings, details

    try:
        with zipfile.ZipFile(file_path, "r") as z:
            all_files = z.namelist()

            # Check for Power Query files
            # Power Query stores queries in xl/queries/ folder
            query_files = [
                f for f in all_files
                if "queries" in f.lower()
                or "query" in f.lower()
            ]

            if query_files:
                for qf in query_files:
                    try:
                        with z.open(qf) as f:
                            content = f.read().decode(
                                "utf-8", errors="ignore"
                            )
                            content_lower = content.lower()

                            # Any HTTP connection in query
                            if ("http://" in content_lower or
                                    "https://" in content_lower):

                                urls = re.findall(
                                    r'https?://[^\s"<>\]]+',
                                    content
                                )
                                for url in urls:
                                    is_safe = any(
                                        s in url.lower()
                                        for s in SAFE_DOMAINS
                                    )
                                    if not is_safe:
                                        findings.append(
                                            "power_query_connection"
                                        )
                                        details.append(
                                            "CRITICAL: Power Query "
                                            "external connection: " +
                                            url
                                        )

                    except Exception:
                        pass

            # Check xl/connections.xml
            if "xl/connections.xml" in all_files:
                with z.open("xl/connections.xml") as f:
                    content = f.read().decode(
                        "utf-8", errors="ignore"
                    )
                    content_lower = content.lower()

                    # OLE DB connections
                    if "oledb" in content_lower:
                        findings.append("ole_db_connection")
                        details.append(
                            "OLE DB connection detected in "
                            "connections.xml"
                        )

                    # Suspicious connection strings
                    suspicious_conn = [
                        "provider=",
                        "data source=http",
                        "extended properties",
                        "exec(",
                        "shell(",
                    ]
                    for sc in suspicious_conn:
                        if sc in content_lower:
                            findings.append(
                                "suspicious_connection"
                            )
                            details.append(
                                "CRITICAL: Suspicious connection "
                                "string: " + sc
                            )

                    # External URLs in connections
                    if ("http://" in content_lower or
                            "https://" in content_lower):
                        urls = re.findall(
                            r'https?://[^\s"<>\]]+', content
                        )
                        for url in urls:
                            is_safe = any(
                                s in url.lower()
                                for s in SAFE_DOMAINS
                            )
                            if not is_safe:
                                findings.append(
                                    "external_data_connection"
                                )
                                details.append(
                                    "External data connection: " + url
                                )

            # Check for external workbook links
            if "xl/externalLinks" in " ".join(all_files):
                ext_links = [
                    f for f in all_files
                    if "externallinks" in f.lower()
                ]
                if ext_links:
                    findings.append("external_data_connection")
                    details.append(
                        "External workbook links found: " +
                        str(len(ext_links)) + " link(s)"
                    )

            if not findings:
                details.append(
                    "No suspicious data connections detected"
                )

    except Exception as e:
        details.append("Power Query check error: " + str(e))

    details.append("Technique 16 findings: " + str(len(findings)))
    return findings, details


# ================================================================
# MAIN FUNCTION — parse_excel
# Called by app.py when user uploads Excel file
# ================================================================

def parse_excel(file_path):
    """
    file_path = full path to uploaded Excel file
    Returns   = complete analysis result dict
    """

    all_findings = []
    all_details  = []
    sha256_hash  = ""

    try:

        all_details.append("=" * 55)
        all_details.append(
            "DARKHOOK_ DEFENCE — EXCEL FILE ANALYSIS"
        )
        all_details.append("=" * 55)
        all_details.append(
            "File: " + os.path.basename(file_path)
        )

        with open(file_path, "rb") as f:
            sha256_hash = hashlib.sha256(f.read()).hexdigest()

        all_details.append("SHA256: " + sha256_hash)
        all_details.append("")

        # Technique 1
        f1, d1 = technique1_file_validation(file_path)
        all_findings.extend(f1)
        all_details.extend(d1)

        # Technique 2
        f2, d2 = technique2_metadata(file_path)
        all_findings.extend(f2)
        all_details.extend(d2)

        # Techniques 3-6
        f3, d3 = techniques3456_macro_analysis(file_path)
        all_findings.extend(f3)
        all_details.extend(d3)

        # Technique 7
        f7, d7 = technique7_embedded_objects(file_path)
        all_findings.extend(f7)
        all_details.extend(d7)

        # Technique 8
        f8, d8 = technique8_external_resources(file_path)
        all_findings.extend(f8)
        all_details.extend(d8)

        # Techniques 9-10
        f9, d9 = techniques910_content_url(file_path)
        all_findings.extend(f9)
        all_details.extend(d9)

        # Technique 11
        f11, d11 = technique11_attack_chain(all_findings)
        all_findings.extend(f11)
        all_details.extend(d11)

        # Technique 12 — scoring
        score, verdict, severity, breakdown = \
            technique12_scoring(all_findings)

        # Technique 13 — XLM
        f13, d13 = technique13_xlm_macros(file_path)
        all_findings.extend(f13)
        all_details.extend(d13)

        # Technique 14 — Hidden sheets
        f14, d14 = technique14_hidden_sheets(file_path)
        all_findings.extend(f14)
        all_details.extend(d14)

        # Technique 15 — Formula injection
        f15, d15 = technique15_formula_injection(file_path)
        all_findings.extend(f15)
        all_details.extend(d15)

        # Technique 16 — Power Query
        f16, d16 = technique16_power_query(file_path)
        all_findings.extend(f16)
        all_details.extend(d16)

        # Final scoring with ALL findings
        score, verdict, severity, breakdown = \
            technique12_scoring(all_findings)

        all_details.append("")
        all_details.append(
            "--- TECHNIQUE 12: HEURISTIC SCORING ---"
        )
        all_details.append("Total techniques run : 16")
        all_details.append(
            "Total findings       : " + str(len(all_findings))
        )
        all_details.append(
            "Danger score         : " + str(score) + "/100"
        )
        all_details.append("Severity             : " + severity)
        all_details.append("Verdict              : " + verdict)
        all_details.append("")
        all_details.append("Score breakdown:")

        for finding, data in breakdown.items():
            all_details.append(
                "  " + finding.ljust(35) +
                " count=" + str(data["count"]) +
                " score=" + str(data["score"])
            )

    except Exception as error:
        all_details.append("Critical error: " + str(error))

    return {
        "findings" : all_findings,
        "details"  : all_details,
        "sha256"   : sha256_hash,
    }
